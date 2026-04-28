"""
AI매출업 엑셀 출력 3종 생성기
① 원시 데이터 엑셀  - 수집된 전체 리뷰·게시물 원본
② 일별 집계 엑셀    - 날짜별 플랫폼 건수 + 누적 + 요일분석
③ 월별 집계 엑셀    - 월별 집계표 + 차트
"""

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

# ── 브랜드 컬러 ──
COLOR_ACCENT  = "FF00D4FF"   # 청록
COLOR_GREEN   = "FF00E5A0"
COLOR_YELLOW  = "FFFFD166"
COLOR_ORANGE  = "FFFF6B35"
COLOR_PURPLE  = "FF7C3AED"
COLOR_BG      = "FF0A0C10"
COLOR_SURFACE = "FF111318"
COLOR_BORDER  = "FF1E2330"
COLOR_TEXT    = "FFE8EAF0"
COLOR_MUTED   = "FF5A6278"
COLOR_WHITE   = "FFFFFFFF"
COLOR_HEADER  = "FF181C24"


def _hdr_font(bold=True, size=10, color=COLOR_TEXT):
    return Font(name="Noto Sans KR", bold=bold, size=size, color=color)

def _cell_font(bold=False, size=10, color=COLOR_TEXT):
    return Font(name="Noto Sans KR", bold=bold, size=size, color=color)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _border() -> Border:
    s = Side(style="thin", color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")

def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def _apply_header_row(ws, row_idx: int, values: list, col_colors: list = None):
    """헤더 행 스타일 적용"""
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.font = _hdr_font(bold=True, size=9, color=COLOR_TEXT)
        cell.fill = _fill(COLOR_HEADER)
        cell.border = _border()
        cell.alignment = _center()


def _apply_data_row(ws, row_idx: int, values: list, is_alt: bool = False, is_total: bool = False):
    """데이터 행 스타일 적용"""
    bg = "FF151A22" if is_alt else COLOR_SURFACE
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        if is_total:
            cell.font = _hdr_font(bold=True, size=10, color=COLOR_ACCENT)
            cell.fill = _fill("FF0D1117")
        else:
            cell.font = _cell_font(size=10, color=COLOR_TEXT)
            cell.fill = _fill(bg)
        cell.border = _border()
        cell.alignment = _right() if isinstance(val, (int, float)) else _left()


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title_block(ws, title: str, subtitle: str, merchant_name: str, period: str):
    """시트 상단 타이틀 블록"""
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 16

    # 타이틀
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"AI매출업 · {title}"
    c.font = Font(name="Noto Sans KR", bold=True, size=16, color=COLOR_ACCENT)
    c.fill = _fill(COLOR_BG)
    c.alignment = _left()

    # 부제
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"{merchant_name}  |  {period}  |  {subtitle}"
    c.font = Font(name="Noto Sans KR", size=10, color=COLOR_MUTED)
    c.fill = _fill(COLOR_BG)
    c.alignment = _left()

    # 생성일
    ws.merge_cells("A3:H3")
    c = ws["A3"]
    c.value = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font = Font(name="Noto Sans KR", size=9, color=COLOR_MUTED)
    c.fill = _fill(COLOR_BG)
    c.alignment = _left()

    # 빈 구분선
    ws.row_dimensions[4].height = 8
    for col in range(1, 9):
        c = ws.cell(row=4, column=col)
        c.fill = _fill(COLOR_BG)


# ══════════════════════════════════════════
# ① 원시 데이터 엑셀
# ══════════════════════════════════════════
def make_raw_excel(report: dict) -> bytes:
    """수집된 전체 리뷰·게시물 원본 엑셀"""
    wb = Workbook()
    wb.remove(wb.active)

    merchant_name = report.get("merchant_name", "")
    period        = report.get("period_label") or report.get("period", "")

    # ── 시트1: 플레이스 영수증 리뷰 ──
    ws = wb.create_sheet("플레이스_영수증리뷰")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "00D4FF"
    _title_block(ws, "네이버 플레이스 영수증 리뷰", "원시 데이터", merchant_name, period)

    headers = ["날짜", "월", "리뷰 유형", "리뷰 내용 (일부)", "출처"]
    _apply_header_row(ws, 5, headers)
    _set_col_widths(ws, [14, 10, 14, 60, 20])

    receipt_items = report.get("chatgpt_payload", {}).get("daily_stats", [])
    # daily_stats에서 영수증 건수 재구성
    place_raw = report.get("_raw_place_items", [])  # 수집기에서 저장한 원시 데이터

    row_i = 6
    if place_raw:
        for i, item in enumerate(place_raw):
            _apply_data_row(ws, row_i, [
                item.get("date", ""),
                item.get("month_key", ""),
                item.get("source", "영수증"),
                item.get("content", "")[:100] if item.get("content") else "",
                "네이버 플레이스",
            ], is_alt=(i % 2 == 1))
            row_i += 1
    else:
        # daily_stats 기반 대체
        daily = report.get("daily_summary", [])
        for i, day in enumerate(daily):
            for _ in range(day.get("place_receipt", 0)):
                _apply_data_row(ws, row_i, [
                    day["date"], day["date"][:7], "영수증 리뷰", "(본문 수집 필요)", "네이버 플레이스"
                ], is_alt=(i % 2 == 1))
                row_i += 1

    if row_i == 6:
        ws.cell(row=6, column=1, value="수집된 데이터 없음").font = Font(color=COLOR_MUTED)

    # ── 시트2: 플레이스 블로그 리뷰 ──
    ws2 = wb.create_sheet("플레이스_블로그리뷰")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "00E5A0"
    _title_block(ws2, "네이버 플레이스 블로그 리뷰", "원시 데이터", merchant_name, period)

    headers2 = ["날짜", "월", "제목/내용", "출처 URL", "플랫폼"]
    _apply_header_row(ws2, 5, headers2)
    _set_col_widths(ws2, [14, 10, 60, 40, 14])

    row_i2 = 6
    place_blog_raw = report.get("_raw_place_blog_items", [])
    if place_blog_raw:
        for i, item in enumerate(place_blog_raw):
            _apply_data_row(ws2, row_i2, [
                item.get("date", ""), item.get("month_key", ""),
                item.get("content", "")[:80], item.get("url", ""), "플레이스 블로그"
            ], is_alt=(i % 2 == 1))
            row_i2 += 1
    else:
        ws2.cell(row=6, column=1, value="수집된 데이터 없음").font = Font(color=COLOR_MUTED)

    # ── 시트3: 네이버 블로그 ──
    ws3 = wb.create_sheet("네이버_블로그")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = "FFD166"
    _title_block(ws3, "네이버 블로그", "원시 데이터", merchant_name, period)

    headers3 = ["날짜", "월", "제목", "블로그 URL", "검색 키워드"]
    _apply_header_row(ws3, 5, headers3)
    _set_col_widths(ws3, [14, 10, 50, 50, 20])

    blog_raw = report.get("_raw_blog_items", [])
    row_i3 = 6
    if blog_raw:
        for i, item in enumerate(blog_raw):
            _apply_data_row(ws3, row_i3, [
                item.get("date", ""), item.get("month_key", ""),
                item.get("title", "")[:80], item.get("url", ""), item.get("keyword", "")
            ], is_alt=(i % 2 == 1))
            row_i3 += 1
    else:
        ws3.cell(row=6, column=1, value="수집된 데이터 없음").font = Font(color=COLOR_MUTED)

    # ── 시트4: 유튜브 ──
    ws4 = wb.create_sheet("유튜브")
    ws4.sheet_view.showGridLines = False
    ws4.sheet_properties.tabColor = "FF6B35"
    _title_block(ws4, "유튜브", "원시 데이터", merchant_name, period)

    headers4 = ["제목", "채널", "조회수", "업로드 시점", "URL"]
    _apply_header_row(ws4, 5, headers4)
    _set_col_widths(ws4, [50, 24, 14, 18, 50])

    top_videos = report.get("top_videos", [])
    row_i4 = 6
    for i, v in enumerate(top_videos):
        _apply_data_row(ws4, row_i4, [
            v.get("title", ""), v.get("channel", ""),
            v.get("views", 0), v.get("published", ""), v.get("url", "")
        ], is_alt=(i % 2 == 1))
        row_i4 += 1

    if not top_videos:
        ws4.cell(row=6, column=1, value="수집된 데이터 없음").font = Font(color=COLOR_MUTED)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════
# ② 일별 집계 엑셀
# ══════════════════════════════════════════
def make_daily_excel(report: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    merchant_name = report.get("merchant_name", "")
    period        = report.get("period_label") or report.get("period", "")
    daily         = report.get("daily_summary", [])

    # ── 시트1: 일별 전체 ──
    ws = wb.create_sheet("일별_전체")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "00D4FF"
    _title_block(ws, "일별 집계", "날짜별 채널 건수", merchant_name, period)

    headers = ["날짜", "블로그", "인스타그램", "영수증 리뷰", "플레이스 블로그", "유튜브", "합계"]
    _apply_header_row(ws, 5, headers)
    _set_col_widths(ws, [14, 12, 12, 14, 16, 10, 10])

    row_i = 6
    for i, day in enumerate(daily):
        _apply_data_row(ws, row_i, [
            day["date"],
            day.get("blog", 0),
            day.get("instagram", 0),
            day.get("place_receipt", 0),
            day.get("place_blog", 0),
            day.get("youtube", 0),
            day.get("total", 0),
        ], is_alt=(i % 2 == 1))
        row_i += 1

    if daily:
        totals = [
            "합계",
            sum(d.get("blog", 0) for d in daily),
            sum(d.get("instagram", 0) for d in daily),
            sum(d.get("place_receipt", 0) for d in daily),
            sum(d.get("place_blog", 0) for d in daily),
            sum(d.get("youtube", 0) for d in daily),
            sum(d.get("total", 0) for d in daily),
        ]
        _apply_data_row(ws, row_i, totals, is_total=True)

    if not daily:
        ws.cell(row=6, column=1, value="수집 완료 후 일별 데이터가 표시됩니다.").font = Font(color=COLOR_MUTED)

    # ── 시트2: 누적 추이 ──
    ws2 = wb.create_sheet("누적_추이")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "00E5A0"
    _title_block(ws2, "누적 추이", "날짜별 누적 합계", merchant_name, period)

    headers2 = ["날짜", "일별 합계", "누적 합계"]
    _apply_header_row(ws2, 5, headers2)
    _set_col_widths(ws2, [14, 14, 14])

    row_i2 = 6
    for i, day in enumerate(daily):
        _apply_data_row(ws2, row_i2, [
            day["date"], day.get("total", 0), day.get("cumulative", 0)
        ], is_alt=(i % 2 == 1))
        row_i2 += 1

    # 꺾은선 차트 - 누적 추이
    if len(daily) >= 2:
        chart = LineChart()
        chart.title = "누적 언급량 추이"
        chart.style = 10
        chart.y_axis.title = "누적 건수"
        chart.x_axis.title = "날짜"
        chart.height = 14
        chart.width  = 24

        data = Reference(ws2, min_col=3, min_row=5, max_row=row_i2 - 1)
        chart.add_data(data, titles_from_data=True)
        chart.series[0].graphicalProperties.line.solidFill = "00D4FF"
        chart.series[0].graphicalProperties.line.width = 25000

        cats = Reference(ws2, min_col=1, min_row=6, max_row=row_i2 - 1)
        chart.set_categories(cats)
        ws2.add_chart(chart, "E5")

    # ── 시트3: 요일 분석 ──
    ws3 = wb.create_sheet("요일_분석")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = "FFD166"
    _title_block(ws3, "요일별 분석", "요일별 평균 언급량", merchant_name, period)

    DOW = ["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"]
    dow_counts = {d: 0 for d in DOW}
    dow_totals = {d: 0 for d in DOW}

    for day in daily:
        try:
            from datetime import date as dt_date
            d_obj = dt_date.fromisoformat(day["date"])
            dow_name = DOW[d_obj.weekday()]
            dow_counts[dow_name] += 1
            dow_totals[dow_name] += day.get("total", 0)
        except Exception:
            continue

    headers3 = ["요일", "총 언급량", "게시 일수", "일평균 언급량"]
    _apply_header_row(ws3, 5, headers3)
    _set_col_widths(ws3, [14, 14, 12, 16])

    row_i3 = 6
    for i, dow in enumerate(DOW):
        cnt   = dow_counts[dow]
        total = dow_totals[dow]
        avg   = round(total / cnt, 1) if cnt > 0 else 0
        _apply_data_row(ws3, row_i3, [dow, total, cnt, avg], is_alt=(i % 2 == 1))
        row_i3 += 1

    # 막대 차트 - 요일별
    if any(dow_totals.values()):
        chart3 = BarChart()
        chart3.title = "요일별 총 언급량"
        chart3.style = 10
        chart3.type  = "col"
        chart3.height = 14
        chart3.width  = 22

        data3 = Reference(ws3, min_col=2, min_row=5, max_row=row_i3 - 1)
        chart3.add_data(data3, titles_from_data=True)
        cats3 = Reference(ws3, min_col=1, min_row=6, max_row=row_i3 - 1)
        chart3.set_categories(cats3)
        chart3.series[0].graphicalProperties.solidFill = "00D4FF"
        ws3.add_chart(chart3, "F5")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════
# ③ 월별 집계 엑셀
# ══════════════════════════════════════════
def make_monthly_excel(report: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    merchant_name = report.get("merchant_name", "")
    period        = report.get("period_label") or report.get("period", "")
    monthly       = report.get("monthly_summary", [])
    summary       = report.get("summary", {})

    # ── 시트1: 월별 집계 ──
    ws = wb.create_sheet("월별_집계")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "00D4FF"
    _title_block(ws, "월별 집계", "플랫폼별 월별 건수", merchant_name, period)

    headers = ["월", "네이버 블로그", "인스타그램", "영수증 리뷰", "플레이스 블로그", "유튜브", "합계", "전월대비"]
    _apply_header_row(ws, 5, headers)
    _set_col_widths(ws, [14, 14, 12, 14, 16, 10, 10, 12])

    prev_total = 0
    row_i = 6
    for i, row in enumerate(monthly):
        total = row.get("total_count", 0)
        growth = ""
        if prev_total > 0:
            pct = round((total - prev_total) / prev_total * 100, 1)
            growth = f"{'+' if pct >= 0 else ''}{pct}%"
        prev_total = total

        _apply_data_row(ws, row_i, [
            row["month"],
            row.get("blog_count", 0),
            row.get("instagram_count", 0),
            row.get("place_receipt_count", 0),
            row.get("place_blog_count", 0),
            row.get("youtube_count", 0),
            total,
            growth,
        ], is_alt=(i % 2 == 1))
        row_i += 1

    # 합계 행
    if monthly:
        _apply_data_row(ws, row_i, [
            "합계",
            sum(r.get("blog_count", 0) for r in monthly),
            sum(r.get("instagram_count", 0) for r in monthly),
            sum(r.get("place_receipt_count", 0) for r in monthly),
            sum(r.get("place_blog_count", 0) for r in monthly),
            sum(r.get("youtube_count", 0) for r in monthly),
            sum(r.get("total_count", 0) for r in monthly),
            "",
        ], is_total=True)

    # 막대 차트 - 월별 채널별
    if len(monthly) >= 1:
        chart = BarChart()
        chart.title = "월별 채널별 언급량"
        chart.style = 10
        chart.type  = "col"
        chart.grouping = "stacked"
        chart.overlap  = 100
        chart.height   = 16
        chart.width    = 28

        cols_info = [
            (2, "네이버 블로그", "00D4FF"),
            (3, "인스타그램",    "00E5A0"),
            (4, "영수증 리뷰",   "FFD166"),
            (5, "플레이스 블로그", "FF6B35"),
            (6, "유튜브",        "7C3AED"),
        ]
        for col_idx, label, color in cols_info:
            data = Reference(ws, min_col=col_idx, min_row=5, max_row=row_i - 1)
            chart.add_data(data, titles_from_data=True)
            idx = cols_info.index((col_idx, label, color))
            if idx < len(chart.series):
                chart.series[idx].graphicalProperties.solidFill = color

        cats = Reference(ws, min_col=1, min_row=6, max_row=row_i - 1)
        chart.set_categories(cats)
        ws.add_chart(chart, "J5")

    # ── 시트2: 요약 대시보드 ──
    ws2 = wb.create_sheet("요약_대시보드")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "FF6B35"
    _title_block(ws2, "요약 대시보드", "기간 전체 합산", merchant_name, period)

    # KPI 블록
    kpis = [
        ("총 언급 수",       summary.get("total_mentions", 0),        "건"),
        ("네이버 블로그",    summary.get("naver_blog_count", 0),       "건"),
        ("인스타그램",       summary.get("instagram_count", 0),        "건"),
        ("영수증 리뷰",      summary.get("place_receipt_count", 0),    "건"),
        ("플레이스 블로그",  summary.get("place_blog_count", 0),       "건"),
        ("유튜브 영상 수",   summary.get("youtube_count", 0),          "건"),
        ("유튜브 총 조회수", summary.get("youtube_total_views", 0),    "회"),
    ]

    ws2.merge_cells("A5:B5")
    ws2["A5"] = "지표"
    ws2["A5"].font = _hdr_font(size=10)
    ws2["A5"].fill = _fill(COLOR_HEADER)
    ws2["A5"].alignment = _center()

    ws2.merge_cells("C5:D5")
    ws2["C5"] = "수치"
    ws2["C5"].font = _hdr_font(size=10)
    ws2["C5"].fill = _fill(COLOR_HEADER)
    ws2["C5"].alignment = _center()

    ws2["E5"] = "단위"
    ws2["E5"].font = _hdr_font(size=10)
    ws2["E5"].fill = _fill(COLOR_HEADER)
    ws2["E5"].alignment = _center()

    _set_col_widths(ws2, [18, 18, 18, 18, 10, 10])

    for i, (label, val, unit) in enumerate(kpis):
        ri = 6 + i
        ws2.merge_cells(f"A{ri}:B{ri}")
        c1 = ws2[f"A{ri}"]
        c1.value = label
        c1.font  = _cell_font(size=11, color=COLOR_TEXT)
        c1.fill  = _fill("FF151A22" if i % 2 else COLOR_SURFACE)
        c1.border = _border()
        c1.alignment = _left()

        ws2.merge_cells(f"C{ri}:D{ri}")
        c2 = ws2[f"C{ri}"]
        c2.value = val
        c2.font  = Font(name="Noto Sans KR", bold=True, size=14, color=COLOR_ACCENT)
        c2.fill  = _fill("FF151A22" if i % 2 else COLOR_SURFACE)
        c2.border = _border()
        c2.alignment = _right()
        c2.number_format = "#,##0"

        c3 = ws2[f"E{ri}"]
        c3.value = unit
        c3.font  = _cell_font(color=COLOR_MUTED)
        c3.fill  = _fill("FF151A22" if i % 2 else COLOR_SURFACE)
        c3.border = _border()
        c3.alignment = _center()

    # 채널 파이 → 막대 차트로 대체 (openpyxl 파이 차트 색상 제한)
    if monthly:
        chart2 = BarChart()
        chart2.title = "채널별 비중"
        chart2.style = 10
        chart2.type  = "bar"
        chart2.height = 14
        chart2.width  = 20

        ch_labels = ["네이버 블로그", "인스타그램", "영수증 리뷰", "플레이스 블로그", "유튜브"]
        ch_values = [
            summary.get("naver_blog_count", 0),
            summary.get("instagram_count", 0),
            summary.get("place_receipt_count", 0),
            summary.get("place_blog_count", 0),
            summary.get("youtube_count", 0),
        ]

        # 임시 데이터 열 작성 (K열~)
        for j, (lbl, val) in enumerate(zip(ch_labels, ch_values)):
            ws2.cell(row=20 + j, column=11, value=lbl)
            ws2.cell(row=20 + j, column=12, value=val)

        data2 = Reference(ws2, min_col=12, min_row=20, max_row=24)
        chart2.add_data(data2)
        cats2 = Reference(ws2, min_col=11, min_row=20, max_row=24)
        chart2.set_categories(cats2)
        chart2.series[0].graphicalProperties.solidFill = "00D4FF"
        ws2.add_chart(chart2, "G5")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
